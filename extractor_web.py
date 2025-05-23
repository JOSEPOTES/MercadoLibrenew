import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
import time
import re
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import unicodedata

# Configuración editable
CONFIG = {
    'main_url': "https://www.carlider.co/listado/accesorios-vehiculos/repuestos-carros-camionetas/",
    'brand_div_selector': 'div.ui-search-filter-dl.shops__filter-items',
    'brand_h3_text': 'Marca',
    'brand_link_selector': 'li.ui-search-filter-container.shops__container-lists a.ui-search-link',
    'show_more_selector': 'a.ui-search-modal__link',
    'modal_grid_selector': 'div.ui-search-search-modal-grid-columns',
    'modal_brand_link_selector': 'a.ui-search-search-modal-filter.ui-search-link',
    'modal_brand_name_selector': 'span.ui-search-search-modal-filter-name',
    'sidebar_brand_name_selector': 'span.ui-search-filter-name.shops-custom-secondary-font',
    'product_title_selector': '.poly-component__title',
    'product_price_selector': '.andes-money-amount.andes-money-amount--cents-superscript .andes-money-amount__fraction',
    'product_link_selector': '.poly-card__content a.poly-component__title',
    'next_page_selector': 'li.andes-pagination__button--next a',
    'base_url': 'https://www.carlider.co',
}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

class CarliderScraper:
    def __init__(self, config):
        self.config = config
        self.session = requests.Session()
        self.session.headers.update(HEADERS)

    def get_soup(self, url):
        response = self.session.get(url)
        return BeautifulSoup(response.text, 'html.parser')

    def extract_all_brands(self):
        soup = self.get_soup(self.config['main_url'])
        brand_div = None
        for div in soup.select(self.config['brand_div_selector']):
            h3 = div.find('h3')
            if h3 and self.config['brand_h3_text'] in h3.get_text():
                brand_div = div
                break
        if not brand_div:
            return []
        brand_links = brand_div.select(self.config['brand_link_selector'])
        show_more = brand_div.select_one(self.config['show_more_selector'])
        if show_more and 'Mostrar más' in show_more.get_text():
            more_url = show_more['href']
            if not more_url.startswith('http'):
                more_url = self.config['base_url'] + more_url
            more_soup = self.get_soup(more_url)
            grid = more_soup.select_one(self.config['modal_grid_selector'])
            if grid:
                modal_brand_links = grid.select(self.config['modal_brand_link_selector'])
                for link in modal_brand_links:
                    name_span = link.select_one(self.config['modal_brand_name_selector'])
                    if name_span:
                        brand_links.append(link)
        brands = []
        for link in brand_links:
            name_span = link.select_one(self.config['sidebar_brand_name_selector'])
            if not name_span:
                name_span = link.select_one(self.config['modal_brand_name_selector'])
            if name_span:
                brands.append({
                    'name': name_span.get_text(strip=True),
                    'url': link['href'] if link['href'].startswith('http') else self.config['base_url'] + link['href']
                })
        return brands

    def extract_products(self, brand_url, progress_callback=None):
        productos = []
        url = brand_url
        # Obtener el nombre de la marca desde la URL (última parte)
        brand_name = brand_url.split('/')[-1].replace('-', ' ').upper()
        while url:
            soup = self.get_soup(url)
            titles = [t.get_text(strip=True) for t in soup.select(self.config['product_title_selector'])]
            prices = [p.get_text(strip=True) for p in soup.select(self.config['product_price_selector'])]
            links = [l['href'] for l in soup.select(self.config['product_link_selector']) if l.has_attr('href') and '/MCO' in l['href']]
            for title, price, link in zip(titles, prices, links):
                image_url = self.extract_main_image(link)
                extra = self.note_detector(link, brand_name)
                producto = {
                    'nombre': title,
                    'valor': price,
                    'unidades': extra['unidades'],
                    'nota': extra['nota'],
                    'proviene': extra['proviene'],
                    'condicion': extra['condicion'],
                    'descripcion_general': extra['descripcion_general'],
                    'categoria_repuesto': extra['categoria_repuesto'],
                    'link': link,
                    'imagen': image_url
                }
                # Validar si el producto está incompleto
                if not producto['valor'] or not producto['link'] or not producto['imagen']:
                    selenium_result = self.buscar_producto_selenium(title)
                    if selenium_result:
                        producto.update(selenium_result)
                        producto['PRODUCTO NO ENCONTRADO'] = ''
                    else:
                        producto['PRODUCTO NO ENCONTRADO'] = 'SI'
                else:
                    producto['PRODUCTO NO ENCONTRADO'] = ''
                productos.append(producto)
                if progress_callback:
                    progress_callback(f"Producto extraído: {title}")
            next_btn = soup.select_one(self.config['next_page_selector'])
            url = next_btn['href'] if next_btn and next_btn.has_attr('href') else None
            if url and not url.startswith('http'):
                url = self.config['base_url'] + url
            time.sleep(1)
        return productos

    def extract_main_image(self, product_url):
        soup = self.get_soup(product_url)
        img_tag = soup.select_one('figure.ui-pdp-gallery__figure img')
        if img_tag and img_tag.has_attr('src'):
            return img_tag['src']
        return ''

    def extract_available_units(self, product_url):
        soup = self.get_soup(product_url)
        # Busca el span con la clase de cantidad disponible
        available_tag = soup.select_one('span.ui-pdp-buybox__quantity__available')
        if available_tag:
            return available_tag.get_text(strip=True)
        return 'No especificado'

    def extract_breadcrumb(self, soup):
        breadcrumb_items = soup.select('li.andes-breadcrumb__item a.andes-breadcrumb__link')
        if breadcrumb_items:
            textos = [a.get_text(strip=True) for a in breadcrumb_items]
            return ' > '.join(textos)
        return ''

    def extract_full_description(self, soup):
        desc_tag = soup.select_one('.ui-pdp-description__content')
        if desc_tag:
            return desc_tag.get_text(separator="\n", strip=True)
        return "Este producto no tiene descripcion."

    def extract_proviene(self, soup, brand_name):
        desc_tag = soup.select_one('.ui-pdp-description__content')
        if desc_tag:
            desc_text = desc_tag.get_text(separator="\n", strip=True)
            match = re.search(r'Viene de\s*([^\n\r]+)', desc_text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return brand_name

    def note_detector(self, product_url, brand_name=None):
        soup = self.get_soup(product_url)
        resultado = {}

        # 1. Notas del producto
        nota = "Producto Usado sin especificaciones"
        desc_tag = soup.select_one('.ui-pdp-description__content')
        if desc_tag:
            desc_text = desc_tag.get_text(separator="\n", strip=True)
            match = re.search(r'(Nota[s]? del producto|Nota[s]?):\s*(.*)', desc_text, re.IGNORECASE)
            if match:
                texto_nota = match.group(2).strip()
                if texto_nota.lower() == "no aplica" or texto_nota == "":
                    nota = "Producto Usado sin especificaciones"
                else:
                    nota = texto_nota.split('\n')[0].strip()
        resultado['nota'] = nota

        # 2. Número de pieza
        numero_pieza = self.extract_numero_pieza(soup)
        resultado['numero_pieza'] = numero_pieza

        # 3. Condición (nuevo o usado)
        condicion = ""
        cond_tag = soup.select_one('.ui-pdp-subtitle')
        if cond_tag:
            condicion = cond_tag.get_text(strip=True)
        resultado['condicion'] = condicion

        # 4. Unidades disponibles
        unidades = "1"
        unidades_tag = soup.select_one('.ui-pdp-buybox__quantity__available')
        if unidades_tag:
            unidades = unidades_tag.get_text(strip=True)
            match = re.search(r'(\d+)', unidades)
            if match:
                unidades = match.group(1)
            else:
                unidades = "1"
        resultado['unidades'] = unidades

        # 5. Proviene
        resultado['proviene'] = self.extract_proviene(soup, brand_name or "")
        # 6. Descripcion general
        resultado['descripcion_general'] = self.extract_full_description(soup)
        # 7. Categoria del repuesto
        resultado['categoria_repuesto'] = self.extract_breadcrumb(soup)

        return resultado

    def extract_numero_pieza(self, soup):
        # Buscar todas las filas de la tabla
        for row in soup.select('tr.andes-table__row'):
            th = row.find('th')
            value_span = row.find('span', class_='andes-table__column--value')
            if th and value_span:
                if "número de pieza" in th.get_text(strip=True).lower():
                    return value_span.get_text(strip=True)
        # Si no encuentra nada, retorna vacío
        return ""

    # --- Lógica Selenium ---
    def buscar_producto_selenium(self, nombre_producto):
        """
        Busca el producto usando Selenium en modo headless y retorna un diccionario
        con los campos completos si lo encuentra, o None si no lo encuentra.
        """
        def normalizar(texto):
            if not texto:
                return ''
            texto = texto.lower().strip()
            texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
            texto = ''.join(e for e in texto if e.isalnum() or e.isspace())
            return texto

        intentos = 0
        while intentos < 2:
            intentos += 1
            try:
                chrome_options = Options()
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--disable-gpu')
                chrome_options.add_argument('--window-size=1920,1080')
                chrome_options.add_argument('--no-sandbox')
                driver = webdriver.Chrome(options=chrome_options)
                driver.get(self.config['main_url'])
                wait = WebDriverWait(driver, 10)
                # Esperar el formulario y el input
                form = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'form#search-form.nav-search-form')))
                input_box = form.find_element(By.CSS_SELECTOR, "input[type='text']")
                input_box.clear()
                input_box.send_keys(nombre_producto)
                button = form.find_element(By.CSS_SELECTOR, "button[type='submit'].search-button")
                button.click()
                # Esperar los resultados
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.poly-card__content a')))
                productos = driver.find_elements(By.CSS_SELECTOR, '.poly-card__content a')
                nombre_normalizado = normalizar(nombre_producto)
                mejor_match = None
                mejor_score = 0
                for prod in productos:
                    try:
                        titulo = prod.text.strip()
                        titulo_normalizado = normalizar(titulo)
                        # Coincidencia exacta
                        if titulo_normalizado == nombre_normalizado:
                            mejor_match = prod
                            mejor_score = 2
                            break
                        # Coincidencia parcial
                        elif nombre_normalizado in titulo_normalizado or titulo_normalizado in nombre_normalizado:
                            if mejor_score < 1:
                                mejor_match = prod
                                mejor_score = 1
                    except Exception:
                        continue
                if mejor_match:
                    link = mejor_match.get_attribute('href')
                    # Entrar al producto
                    driver.execute_script("window.open(arguments[0]);", link)
                    driver.switch_to.window(driver.window_handles[-1])
                    try:
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.andes-money-amount__fraction')))
                    except TimeoutException:
                        print(f"[Selenium] Timeout esperando detalles para: {nombre_producto}")
                    # Extraer datos igual que en el scraping normal
                    try:
                        precio = driver.find_element(By.CSS_SELECTOR, '.andes-money-amount.andes-money-amount--cents-superscript .andes-money-amount__fraction').text.strip()
                    except NoSuchElementException:
                        precio = ''
                    try:
                        imagen = driver.find_element(By.CSS_SELECTOR, 'figure.ui-pdp-gallery__figure img').get_attribute('src')
                    except NoSuchElementException:
                        imagen = ''
                    try:
                        nota = driver.find_element(By.CSS_SELECTOR, '.ui-pdp-description__content').text.strip()
                    except NoSuchElementException:
                        nota = ''
                    try:
                        condicion = driver.find_element(By.CSS_SELECTOR, '.ui-pdp-subtitle').text.strip()
                    except NoSuchElementException:
                        condicion = ''
                    try:
                        unidades = driver.find_element(By.CSS_SELECTOR, '.ui-pdp-buybox__quantity__available').text.strip()
                    except NoSuchElementException:
                        unidades = '1'
                    # Número de pieza (opcional)
                    numero_pieza = ''
                    try:
                        filas = driver.find_elements(By.CSS_SELECTOR, 'tr.andes-table__row')
                        for row in filas:
                            th = row.find_element(By.TAG_NAME, 'th').text.strip().lower()
                            if 'número de pieza' in th:
                                numero_pieza = row.find_element(By.CSS_SELECTOR, 'span.andes-table__column--value').text.strip()
                                break
                    except Exception:
                        numero_pieza = ''
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.quit()
                    return {
                        'valor': precio,
                        'link': link,
                        'imagen': imagen,
                        'nota': nota,
                        'numero_pieza': numero_pieza,
                        'condicion': condicion,
                        'unidades': unidades
                    }
                else:
                    print(f"[Selenium] No se encontró coincidencia para: {nombre_producto}")
                driver.quit()
            except Exception as e:
                print(f"[Selenium] Error buscando '{nombre_producto}': {e}")
                try:
                    driver.quit()
                except:
                    pass
        return None

class App(tk.Tk):
    def __init__(self, scraper):
        super().__init__()
        self.scraper = scraper
        self.title("Extractor Carlider Futurista")
        self.geometry("540x500")
        self.configure(bg="#181c24")
        self.resizable(False, False)
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.style.configure('TButton', font=('Segoe UI', 12, 'bold'), foreground='#fff', background='#00b4d8', borderwidth=0, focusthickness=3, focuscolor='none')
        self.style.map('TButton', background=[('active', '#0077b6')])
        self.style.configure('TLabel', background="#181c24", foreground="#fff", font=('Segoe UI', 11))
        self.style.configure('TCombobox', fieldbackground="#23293a", background="#23293a", foreground="#fff", font=('Segoe UI', 11))
        self.create_widgets()
        self.brands = []
        self.load_brands()

    def create_widgets(self):
        self.title_label = tk.Label(self, text="EXTRACTOR CARLIDER", font=("Orbitron", 20, "bold"), fg="#00b4d8", bg="#181c24")
        self.title_label.pack(pady=(18, 8))
        self.label = ttk.Label(self, text="Selecciona la marca o 'Todos':")
        self.label.pack(pady=5)
        self.combo = ttk.Combobox(self, state="readonly", width=30)
        self.combo.pack(pady=5)
        self.combo['values'] = ["Cargando..."]
        self.combo.current(0)
        self.btn = ttk.Button(self, text="⬇ Descargar", command=self.on_download)
        self.btn.pack(pady=15, ipadx=10, ipady=5)
        self.text_frame = tk.Frame(self, bg="#23293a", bd=2, relief="groove")
        self.text_frame.pack(padx=15, pady=10, fill=tk.BOTH, expand=True)
        self.text = tk.Text(self.text_frame, height=15, width=60, bg="#23293a", fg="#fff", font=("Consolas", 10), bd=0, relief="flat", wrap=tk.WORD)
        self.text.pack(padx=8, pady=8, fill=tk.BOTH, expand=True)
        self.text.config(state=tk.DISABLED)

    def load_brands(self):
        self.text.config(state=tk.NORMAL)
        self.text.insert(tk.END, "Cargando marcas, espera...\n")
        self.text.config(state=tk.DISABLED)
        self.update()
        self.brands = self.scraper.extract_all_brands()
        unique = {}
        for b in self.brands:
            if b['name'] not in unique:
                unique[b['name']] = b
        self.brands = list(unique.values())
        brand_names = [b['name'] for b in self.brands]
        self.combo['values'] = ["Todos"] + brand_names
        self.combo.current(0)
        self.text.config(state=tk.NORMAL)
        self.text.insert(tk.END, "Marcas cargadas.\n")
        self.text.config(state=tk.DISABLED)

    def on_download(self):
        selected = self.combo.get()
        if selected == "Cargando...":
            messagebox.showinfo("Espera", "Las marcas aún se están cargando.")
            return
        self.text.config(state=tk.NORMAL)
        self.text.delete(1.0, tk.END)
        if selected == "Todos":
            self.text.insert(tk.END, "Descargando todas las marcas...\n")
            self.download_all_brands()
        else:
            self.text.insert(tk.END, f"Descargando productos de {selected}...\n")
            self.download_brand(selected)
        self.text.config(state=tk.DISABLED)

    def progress(self, msg):
        self.text.config(state=tk.NORMAL)
        self.text.insert(tk.END, msg + "\n")
        self.text.see(tk.END)
        self.text.config(state=tk.DISABLED)
        self.update()

    def download_brand(self, brand_name):
        brand = next((b for b in self.brands if b['name'] == brand_name), None)
        if not brand:
            self.progress(f"No se encontró la marca: {brand_name}")
            return
        productos = self.scraper.extract_products(brand['url'], self.progress)
        clean_name = re.sub(r'[^a-z0-9]', '', brand_name.lower().replace(' ', ''))
        output_dir = r'C:\Users\Joseg\Desktop\Carlider\ExtractorCarlider'
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, f'{clean_name}.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = brand_name
        ws.append(['Nombre', 'Valor', 'Unidades', 'Notas del producto', 'Proviene', 'Condicion', 'Descripcion general', 'Categoria del Repuesto', 'Link', 'Imagen'])
        for p in productos:
            ws.append([
                p.get('nombre', ''),
                p.get('valor', ''),
                p.get('unidades', ''),
                p.get('nota', ''),
                p.get('proviene', ''),
                p.get('condicion', ''),
                p.get('descripcion_general', ''),
                p.get('categoria_repuesto', ''),
                p.get('link', ''),
                p.get('imagen', '')
            ])
        wb.save(filename)
        self.progress(f"Archivo guardado: {filename}")

    def download_all_brands(self):
        wb = Workbook()
        wb.remove(wb.active)
        sheet_names = set()
        for brand in self.brands:
            productos = self.scraper.extract_products(brand['url'], lambda msg: self.progress(f"[{brand['name']}] {msg}"))
            sheet_name = brand['name'][:31]
            original_name = sheet_name
            i = 1
            while sheet_name in sheet_names:
                sheet_name = f"{original_name[:28]}_{i}"
                i += 1
            sheet_names.add(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Nombre', 'Valor', 'Unidades', 'Notas del producto', 'Proviene', 'Condicion', 'Descripcion general', 'Categoria del Repuesto', 'Link', 'Imagen'])
            for p in productos:
                ws.append([
                    p.get('nombre', ''),
                    p.get('valor', ''),
                    p.get('unidades', ''),
                    p.get('nota', ''),
                    p.get('proviene', ''),
                    p.get('condicion', ''),
                    p.get('descripcion_general', ''),
                    p.get('categoria_repuesto', ''),
                    p.get('link', ''),
                    p.get('imagen', '')
                ])
        output_dir = r'C:\Users\Joseg\Desktop\Carlider\ExtractorCarlider'
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, 'todas_las_marcas.xlsx')
        wb.save(filename)
        self.progress(f"Archivo guardado: {filename}")

if __name__ == '__main__':
    scraper = CarliderScraper(CONFIG)
    app = App(scraper)
    app.mainloop() 